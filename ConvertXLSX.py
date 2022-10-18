# -*- coding: utf-8 -*-
"""
Created on Thu Feb 17 17:15:17 2022

@author: mwodring
"""

# -*- coding: utf-8 -*-
# Script to extract QuantStudio 6&7 exports to csv files then create 
# plots from these using an R Script.
# Adapted in part from: 
# https://yagisanatode.com/2017/11/18/copy-and-paste-ranges-in-excel-with-openpyxl-and-python-3/

from openpyxl import load_workbook
from openpyxl import Workbook
import os
import csv

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
 
    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def excelSlicer(filepath):
    workbook = load_workbook(filename = filepath)
    sheet = workbook.active
    selectedRange = copyRange(1,45,9,8335, sheet)
    newWorkbook = Workbook()
    slicedSheet = newWorkbook.active
    pasteRange(1,1,9,3841,slicedSheet,selectedRange)
    newfilename = filepath[0:-5] + "_sliced" + ".xlsx"
    newWorkbook.save(filename = newfilename)
    return newfilename

def spacesToDots(sheetin):
    sheet = sheetin
    for row in sheet.iter_rows(min_row=1, 
                                 max_row=1, 
                                 min_col=1, 
                                 max_col = 9):
        for cell in row:
            cellstr = str(cell.value)
            celltoconcat = []
            for letter in cellstr:
                if letter == " ":
                    celltoconcat.append(".")
                else:
                    celltoconcat.append(letter)
            newname = "".join(celltoconcat)
            sheet[cell.coordinate].value = newname
            
                
                
def toCSV(filein, fileout):
    workbook = load_workbook(filein)
    sheet = workbook.active
    spacesToDots(sheet)
    with open(fileout, 'w', newline="") as file:
        newcsv = csv.writer(file)
        for row in sheet.rows:
            newcsv.writerow([cell.value for cell in row])

def convertAllFiles(path):
    csvlist = []
    excelfiles = []
    with os.scandir(path) as folder:
        for file in folder:
            if file.name[-11:-5] == "sliced" and file.name[-4:] == ".csv":
                csvlist.append(file)
            else:
                if file.name.endswith(".xlsx") and file.is_file() and not file.name[-11:-5] == "sliced":
                    excelfiles.append(file.name)        
    
    filestoconvert = []
    
    for filename in excelfiles:
        if filename[0] != "$":
            filestoconvert.append(excelSlicer(path + "\\" + filename))
    
    for file in filestoconvert:
          csvname = file[:-5] + ".csv"
          csvlist.append(csvname)
          toCSV(file, csvname)
        
    return csvlist
